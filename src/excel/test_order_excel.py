import pdb

from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader

file_path = './order.xlsx'


if __name__ == '__main__':
    wb = load_workbook(file_path)
    sheet = wb.active
    image_loader = SheetImageLoader(sheet)
    cell = sheet.cell(7, 2)
    image = image_loader.get('B8')
    # pdb.set_trace()

    image = image_loader.get('D7')
    cell = sheet.cell(7, 4)
    pdb.set_trace()
